#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import os
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


@dataclass(frozen=True)
class RowOp:
    excel_path: Path
    sheet_name: str
    excel_row_number: int  # 1-based, includes header row (so data starts at 2)
    data_row_index: int  # 1-based within non-empty audio_file rows
    old_cell_value: str
    old_abs_path: Path
    new_abs_path: Path
    new_cell_value: str


def _norm_rel_path(cell_value: str) -> str:
    v = (cell_value or "").strip().strip('"').strip("'")
    # Normalize Windows/Unix separators to forward slash for storage in Excel
    v = v.replace("\\", "/")
    # Collapse duplicate slashes (avoid turning "http://" etc; not relevant here)
    while "//" in v:
        v = v.replace("//", "/")
    return v


def _resolve_old_abs_path(collected_dir: Path, cell_value: str) -> Tuple[Path, str]:
    """
    Returns (abs_path, normalized_cell_value_used_for_resolution)

    If cell value ends with .flac, treat as relative path.
    Else treat it as an ID and resolve to cleaned_audio/<id>.flac.
    """
    v = _norm_rel_path(cell_value)
    if not v:
        raise ValueError("empty audio_file cell")

    if v.lower().endswith(".flac"):
        rel = v
    else:
        rel = f"cleaned_audio/{v}.flac"

    abs_path = (collected_dir / rel).resolve()
    return abs_path, rel


def _ensure_inside(parent: Path, child: Path) -> None:
    parent_r = parent.resolve()
    child_r = child.resolve()
    try:
        child_r.relative_to(parent_r)
    except ValueError as e:
        raise ValueError(f"path escapes expected directory: {child_r} not in {parent_r}") from e


def _find_header_col(ws, header_name: str) -> int:
    header_name_norm = header_name.strip().lower()
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip().lower() == header_name_norm:
            return c
    raise KeyError(f"header '{header_name}' not found in row 1")


def _iter_row_ops_for_excel(
    collected_dir: Path, excel_path: Path, sheet_name: str = "Sheet1"
) -> List[RowOp]:
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"{excel_path.name}: sheet '{sheet_name}' not found (have {wb.sheetnames})")

    ws = wb[sheet_name]
    audio_col = _find_header_col(ws, "audio_file")

    cleaned_audio_dir = (collected_dir / "cleaned_audio").resolve()
    _ensure_inside(collected_dir, cleaned_audio_dir)

    ops: List[RowOp] = []
    data_row_index = 0
    stem = excel_path.stem

    # Iterate through all used rows; skip blanks in audio_file
    for r in range(2, (ws.max_row or 1) + 1):
        cell = ws.cell(row=r, column=audio_col)
        raw = cell.value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            continue
        data_row_index += 1

        raw_str = str(raw)
        old_abs, _old_rel_used = _resolve_old_abs_path(collected_dir, raw_str)
        _ensure_inside(cleaned_audio_dir, old_abs)

        new_name = f"{stem}{data_row_index}.flac"
        new_abs = (cleaned_audio_dir / new_name).resolve()
        _ensure_inside(cleaned_audio_dir, new_abs)

        ops.append(
            RowOp(
                excel_path=excel_path,
                sheet_name=sheet_name,
                excel_row_number=r,
                data_row_index=data_row_index,
                old_cell_value=raw_str,
                old_abs_path=old_abs,
                new_abs_path=new_abs,
                new_cell_value=f"cleaned_audio/{new_name}",
            )
        )

    return ops


def _validate_ops(all_ops: List[RowOp]) -> None:
    if not all_ops:
        raise ValueError("no operations generated (no non-empty audio_file rows?)")

    # Each old file should be referenced once across scope (fail fast if not)
    old_map: Dict[Path, List[RowOp]] = {}
    for op in all_ops:
        old_map.setdefault(op.old_abs_path, []).append(op)
    multi_old = [(p, v) for p, v in old_map.items() if len(v) > 1]
    if multi_old:
        example = multi_old[0]
        raise ValueError(
            f"same source file referenced multiple times: {example[0]} ({len(example[1])} rows); "
            f"refuse to proceed"
        )

    # Destinations must be unique
    new_map: Dict[Path, List[RowOp]] = {}
    for op in all_ops:
        new_map.setdefault(op.new_abs_path, []).append(op)
    multi_new = [(p, v) for p, v in new_map.items() if len(v) > 1]
    if multi_new:
        example = multi_new[0]
        raise ValueError(
            f"destination collision: {example[0]} would be written by {len(example[1])} rows; refuse to proceed"
        )

    # Existence checks (source must exist; destination must not exist unless same path)
    missing_sources = [op for op in all_ops if not op.old_abs_path.exists()]
    if missing_sources:
        ex = missing_sources[0]
        raise FileNotFoundError(f"missing source file: {ex.old_abs_path} (from {ex.excel_path.name} row {ex.excel_row_number})")

    old_paths = set(old_map.keys())
    # Destination can exist if it's also a source path in this same batch (we'll use a temp-rename strategy).
    existing_dests = [
        op
        for op in all_ops
        if op.new_abs_path.exists()
        and op.new_abs_path != op.old_abs_path
        and op.new_abs_path not in old_paths
    ]
    if existing_dests:
        ex = existing_dests[0]
        raise FileExistsError(
            f"destination already exists: {ex.new_abs_path} (from {ex.excel_path.name} row {ex.excel_row_number}); "
            f"refuse to proceed"
        )


def _write_report(report_path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows)
    fieldnames: List[str] = [
        "excel_file",
        "sheet",
        "data_row_index",
        "excel_row_number",
        "old_cell_value",
        "old_abs_path",
        "new_cell_value",
        "new_abs_path",
        "action",
        "status",
        "error",
    ]
    with report_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})


def _apply_ops(
    collected_dir: Path,
    all_ops: List[RowOp],
    dry_run: bool,
    report_path: Path,
) -> None:
    # Group ops by Excel file for efficient workbook updates
    by_excel: Dict[Path, List[RowOp]] = {}
    for op in all_ops:
        by_excel.setdefault(op.excel_path, []).append(op)

    report_rows: List[Dict[str, Any]] = []

    rename_ops = [op for op in all_ops if op.old_abs_path != op.new_abs_path]
    temp_for_old: Dict[Path, Path] = {}

    def _make_temp_path(op: RowOp) -> Path:
        # Short uuid suffix avoids collisions while keeping names reasonable.
        suffix = uuid.uuid4().hex[:12]
        name = f".tmp_rename_{op.excel_path.stem}_{op.data_row_index}_{suffix}.flac"
        return op.old_abs_path.with_name(name)

    # First: file renames (so Excel points at final names)
    # Use two-phase rename to avoid collisions when destination names already exist as sources.
    if dry_run:
        for op in all_ops:
            report_rows.append(
                {
                    "excel_file": op.excel_path.name,
                    "sheet": op.sheet_name,
                    "data_row_index": op.data_row_index,
                    "excel_row_number": op.excel_row_number,
                    "old_cell_value": op.old_cell_value,
                    "old_abs_path": str(op.old_abs_path),
                    "new_cell_value": op.new_cell_value,
                    "new_abs_path": str(op.new_abs_path),
                    "action": "skip_rename_same_path" if op.old_abs_path == op.new_abs_path else "rename_via_temp",
                    "status": "ok",
                    "error": "",
                }
            )
    else:
        # Phase 1: move every renamed source to a temp name
        for op in rename_ops:
            row: Dict[str, Any] = {
                "excel_file": op.excel_path.name,
                "sheet": op.sheet_name,
                "data_row_index": op.data_row_index,
                "excel_row_number": op.excel_row_number,
                "old_cell_value": op.old_cell_value,
                "old_abs_path": str(op.old_abs_path),
                "new_cell_value": op.new_cell_value,
                "new_abs_path": str(op.new_abs_path),
                "action": "rename_via_temp",
            }
            try:
                tmp = _make_temp_path(op)
                if tmp.exists():
                    raise FileExistsError(f"temp path already exists: {tmp}")
                os.replace(op.old_abs_path, tmp)
                temp_for_old[op.old_abs_path] = tmp
                row["status"] = "ok"
            except Exception as e:
                row["status"] = "error"
                row["error"] = repr(e)
                report_rows.append(row)
                _write_report(report_path, report_rows)
                raise
            report_rows.append(row)

        # Phase 2: move temp files to final destinations
        for op in rename_ops:
            row: Dict[str, Any] = {
                "excel_file": op.excel_path.name,
                "sheet": op.sheet_name,
                "data_row_index": op.data_row_index,
                "excel_row_number": op.excel_row_number,
                "old_cell_value": op.old_cell_value,
                "old_abs_path": str(op.old_abs_path),
                "new_cell_value": op.new_cell_value,
                "new_abs_path": str(op.new_abs_path),
                "action": "finalize_from_temp",
            }
            try:
                tmp = temp_for_old[op.old_abs_path]
                os.replace(tmp, op.new_abs_path)
                row["status"] = "ok"
            except Exception as e:
                row["status"] = "error"
                row["error"] = repr(e)
                report_rows.append(row)
                _write_report(report_path, report_rows)
                raise
            report_rows.append(row)

        # Also record the no-op rows for completeness
        for op in all_ops:
            if op.old_abs_path != op.new_abs_path:
                continue
            report_rows.append(
                {
                    "excel_file": op.excel_path.name,
                    "sheet": op.sheet_name,
                    "data_row_index": op.data_row_index,
                    "excel_row_number": op.excel_row_number,
                    "old_cell_value": op.old_cell_value,
                    "old_abs_path": str(op.old_abs_path),
                    "new_cell_value": op.new_cell_value,
                    "new_abs_path": str(op.new_abs_path),
                    "action": "skip_rename_same_path",
                    "status": "ok",
                    "error": "",
                }
            )

    # Second: Excel updates (only audio_file column)
    for excel_path, ops in by_excel.items():
        try:
            wb = load_workbook(excel_path)
            ws = wb[ops[0].sheet_name]
            audio_col = _find_header_col(ws, "audio_file")

            for op in ops:
                ws.cell(row=op.excel_row_number, column=audio_col).value = op.new_cell_value

            if not dry_run:
                wb.save(excel_path)
        except Exception as e:
            _write_report(report_path, report_rows)
            raise RuntimeError(f"failed updating {excel_path.name}: {e}") from e

    _write_report(report_path, report_rows)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Rename cleaned_audio FLACs based on Collected/A.xlsx..E.xlsx row order and update audio_file column.",
    )
    ap.add_argument(
        "--collected-dir",
        default="Collected",
        help="Path to Collected directory (default: Collected).",
    )
    ap.add_argument(
        "--excels",
        nargs="*",
        default=["A.xlsx", "B.xlsx", "C.xlsx", "D.xlsx", "E.xlsx"],
        help="Excel filenames (relative to collected-dir) to process (default: A.xlsx..E.xlsx).",
    )
    ap.add_argument(
        "--sheet",
        default="Sheet1",
        help="Sheet name to use (default: Sheet1).",
    )
    ap.add_argument(
        "--report",
        default="rename_report.csv",
        help="Report CSV filename (relative to collected-dir unless absolute).",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not rename or save Excel; only validate and write the report.",
    )
    args = ap.parse_args()

    collected_dir = Path(args.collected_dir).expanduser()
    if not collected_dir.is_absolute():
        collected_dir = (Path.cwd() / collected_dir).resolve()

    excel_paths = [collected_dir / x for x in args.excels]
    for p in excel_paths:
        if not p.exists():
            raise FileNotFoundError(f"excel not found: {p}")

    report_path = Path(args.report).expanduser()
    if not report_path.is_absolute():
        report_path = (collected_dir / report_path).resolve()

    all_ops: List[RowOp] = []
    for ep in excel_paths:
        all_ops.extend(_iter_row_ops_for_excel(collected_dir, ep, sheet_name=args.sheet))

    _validate_ops(all_ops)

    total_rows = len(all_ops)
    print(f"Prepared {total_rows} row operations across {len(excel_paths)} Excel files.")
    print(f"Report will be written to: {report_path}")
    if args.dry_run:
        print("Dry-run enabled: no files will be renamed and Excel files will not be saved.")

    _apply_ops(collected_dir, all_ops, dry_run=args.dry_run, report_path=report_path)

    if args.dry_run:
        print("Dry-run complete.")
    else:
        print("Rename + Excel update complete.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

