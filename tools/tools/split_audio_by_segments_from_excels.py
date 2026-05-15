#!/usr/bin/env python3
from __future__ import annotations

import argparse
import ast
import json
import math
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import soundfile as sf
from openpyxl import load_workbook


@dataclass(frozen=True)
class Segment:
    start_sec: float
    end_sec: float
    text: str


def _norm_rel_path(cell_value: str) -> str:
    v = (cell_value or "").strip().strip('"').strip("'")
    v = v.replace("\\", "/")
    while "//" in v:
        v = v.replace("//", "/")
    return v


def _safe_name(s: str, max_len: int = 80) -> str:
    s = (s or "").strip()
    if not s:
        return "unknown"
    s = re.sub(r"[^\w\-.]+", "_", s, flags=re.UNICODE).strip("_")
    return s[:max_len] if len(s) > max_len else s


def _find_header_map(ws) -> Dict[str, int]:
    max_col = ws.max_column or 0
    header_map: Dict[str, int] = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str):
            key = v.strip().lower()
            if key:
                header_map[key] = c
    return header_map


def _parse_segments_cell(raw: Any) -> Tuple[List[Dict[str, Any]], str]:
    """
    Returns (segments_list, parse_kind)

    parse_kind is one of:
      - "native" (already list)
      - "ast" (parsed via ast.literal_eval)
      - "json" (parsed via json.loads)
    """
    if raw is None:
        return [], "native"
    if isinstance(raw, list):
        return raw, "native"
    if isinstance(raw, (dict, tuple)):
        return [dict(raw)], "native"
    if not isinstance(raw, str):
        raise ValueError(f"unsupported segments cell type: {type(raw)}")

    s = raw.strip()
    if not s:
        return [], "native"

    try:
        v = ast.literal_eval(s)
        if isinstance(v, list):
            return v, "ast"
        if isinstance(v, dict):
            return [v], "ast"
    except Exception:
        pass

    v = json.loads(s)
    if isinstance(v, list):
        return v, "json"
    if isinstance(v, dict):
        return [v], "json"
    raise ValueError("segments parsed to unexpected type")


def _coerce_segment(item: Dict[str, Any]) -> Optional[Segment]:
    if not isinstance(item, dict):
        return None
    start = item.get("start", None)
    end = item.get("end", None)
    text = item.get("text", "") or ""
    try:
        start_f = float(start)
        end_f = float(end)
    except Exception:
        return None
    if not math.isfinite(start_f) or not math.isfinite(end_f):
        return None
    if end_f <= start_f:
        return None
    return Segment(start_sec=start_f, end_sec=end_f, text=str(text).strip())


def _iter_excels(collected_dir: Path) -> List[Path]:
    return sorted(
        p.resolve()
        for p in collected_dir.glob("*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    )


def _is_generated_manifest_excel(path: Path) -> bool:
    name = path.name.lower()
    return name.startswith("segmented_dataset") and name.endswith(".xlsx")


def _write_manifest_excel(df: pd.DataFrame, output_excel_path: Path) -> Path:
    """
    Write df to output_excel_path. Uses a temp file + replace to reduce partial writes.
    Returns the path written (may differ if the target is locked).
    """
    output_excel_path.parent.mkdir(parents=True, exist_ok=True)

    tmp = output_excel_path.with_name(output_excel_path.stem + ".tmp" + output_excel_path.suffix)
    try:
        df.to_excel(tmp, index=False, engine="openpyxl")
        tmp.replace(output_excel_path)
        return output_excel_path
    except PermissionError:
        ts = time.strftime("%Y%m%d_%H%M%S")
        alt = output_excel_path.with_name(f"{output_excel_path.stem}_{ts}{output_excel_path.suffix}")
        df.to_excel(alt, index=False, engine="openpyxl")
        return alt
    finally:
        try:
            if tmp.exists():
                tmp.unlink()
        except OSError:
            pass


def _resolve_audio_path(collected_dir: Path, audio_cell: Any) -> Tuple[Path, str]:
    if audio_cell is None:
        raise ValueError("empty audio_file cell")
    raw = str(audio_cell).strip()
    if not raw:
        raise ValueError("empty audio_file cell")

    raw_norm = _norm_rel_path(raw)
    p = Path(raw_norm)
    if p.is_absolute() or (len(p.drive) > 0 and ":" in raw_norm[:3]):
        return Path(raw), raw

    abs_path = (collected_dir / raw_norm).resolve()
    return abs_path, raw_norm


def _read_audio_clip(audio_path: Path, start_sec: float, end_sec: float) -> Tuple[Any, int]:
    with sf.SoundFile(str(audio_path)) as f:
        sr = int(f.samplerate)
        total = int(f.frames)
        start_frame = int(round(start_sec * sr))
        end_frame = int(round(end_sec * sr))
        start_frame = max(0, min(total, start_frame))
        end_frame = max(0, min(total, end_frame))
        if end_frame <= start_frame:
            raise ValueError("empty clip after clamping")
        f.seek(start_frame)
        data = f.read(end_frame - start_frame, dtype="float32", always_2d=True)
    return data, sr


def _write_flac_if_missing(path: Path, data: Any, sr: int) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    sf.write(str(path), data, sr, format="FLAC")


def run(
    collected_dir: Path,
    output_audio_dir: Path,
    output_excel_path: Path,
    inspect_only: bool,
    max_rows: int,
    verbose: bool,
) -> int:
    excels = _iter_excels(collected_dir)
    # Prevent reading our own generated manifests as input (and avoid self-locking).
    excels = [p for p in excels if not _is_generated_manifest_excel(p)]
    if not excels:
        print(f"No .xlsx found in {collected_dir}")
        return 2

    manifest_rows: List[Dict[str, Any]] = []
    parse_kind_counts: Dict[str, int] = {}
    missing_required_sheets: List[str] = []
    errors: List[str] = []

    for excel_path in excels:
        wb = None
        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            if verbose or inspect_only:
                print(f"\n=== {excel_path.name} ({len(sheet_names)} sheet(s)) ===")

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                headers = _find_header_map(ws)
                if "audio_file" not in headers or "segments" not in headers:
                    missing_required_sheets.append(f"{excel_path.name}:{sheet_name}")
                    if verbose or inspect_only:
                        print(
                            f"- skip sheet {sheet_name!r}: missing required headers (have: {sorted(headers.keys())})"
                        )
                    continue

                audio_col = headers["audio_file"]
                seg_col = headers["segments"]

                optional_cols = {
                    k: v
                    for k, v in headers.items()
                    if k not in {"audio_file", "segments"}
                }

                n_processed_rows = 0
                n_segments_total = 0
                n_segments_written = 0

                for r in range(2, (ws.max_row or 1) + 1):
                    if max_rows > 0 and n_processed_rows >= max_rows:
                        break

                    audio_cell = ws.cell(row=r, column=audio_col).value
                    seg_cell = ws.cell(row=r, column=seg_col).value
                    if audio_cell is None or (isinstance(audio_cell, str) and not audio_cell.strip()):
                        continue
                    if seg_cell is None or (isinstance(seg_cell, str) and not seg_cell.strip()):
                        continue

                    n_processed_rows += 1

                    try:
                        audio_abs, audio_rel_norm = _resolve_audio_path(collected_dir, audio_cell)
                    except Exception as e:
                        errors.append(f"{excel_path.name}:{sheet_name}:row{r}: bad audio_file: {e}")
                        continue

                    if not audio_abs.exists():
                        errors.append(
                            f"{excel_path.name}:{sheet_name}:row{r}: missing audio file: {audio_abs}"
                        )
                        continue

                    try:
                        seg_list, kind = _parse_segments_cell(seg_cell)
                        parse_kind_counts[kind] = parse_kind_counts.get(kind, 0) + 1
                    except Exception as e:
                        errors.append(f"{excel_path.name}:{sheet_name}:row{r}: segments parse failed: {e}")
                        continue

                    segments: List[Segment] = []
                    for item in seg_list:
                        seg = _coerce_segment(item) if isinstance(item, dict) else None
                        if seg is not None:
                            segments.append(seg)
                    if not segments:
                        errors.append(f"{excel_path.name}:{sheet_name}:row{r}: no valid segments after validation")
                        continue

                    n_segments_total += len(segments)

                    # Pull optional columns from the row for manifest
                    optional_values: Dict[str, Any] = {}
                    for col_name, cidx in optional_cols.items():
                        optional_values[col_name] = ws.cell(row=r, column=cidx).value

                    for idx, seg in enumerate(segments):
                        # One folder per Excel stem; collision-safe filename.
                        # Example: segmented_audio/A/Sheet1_row_2_seg_000.flac
                        out_rel = (
                            Path(output_audio_dir.name)
                            / _safe_name(excel_path.stem)
                            / f"{_safe_name(sheet_name)}_row_{r}_seg_{idx:03d}.flac"
                        )
                        out_abs = (collected_dir / out_rel).resolve()

                        if not inspect_only:
                            try:
                                data, sr = _read_audio_clip(audio_abs, seg.start_sec, seg.end_sec)
                                _write_flac_if_missing(out_abs, data, sr)
                                n_segments_written += 1
                            except Exception as e:
                                errors.append(
                                    f"{excel_path.name}:{sheet_name}:row{r}:seg{idx}: audio split failed: {e}"
                                )
                                continue

                        manifest_rows.append(
                            {
                                "source_excel": excel_path.name,
                                "source_sheet": sheet_name,
                                "source_excel_row": r,
                                "source_audio_file": audio_rel_norm,
                                "segment_index": idx,
                                "start_sec": float(seg.start_sec),
                                "end_sec": float(seg.end_sec),
                                "duration_sec": float(seg.end_sec - seg.start_sec),
                                "segment_audio_file": str(out_rel).replace("\\", "/"),
                                "segment_transcription": seg.text,
                                **optional_values,
                            }
                        )

                if verbose or inspect_only:
                    print(
                        f"- sheet {sheet_name!r}: rows={n_processed_rows}, segments={n_segments_total}"
                        + ("" if inspect_only else f", written={n_segments_written}")
                    )
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    if inspect_only:
        print("\n--- Inspect summary ---")
        print(f"excels_found: {len(excels)}")
        print(f"parse_kinds: {parse_kind_counts}")
        if missing_required_sheets:
            print(f"sheets_missing_required_headers: {len(missing_required_sheets)}")
            if verbose:
                for x in missing_required_sheets[:50]:
                    print("  -", x)
        if errors:
            print(f"errors: {len(errors)} (showing first 20)")
            for e in errors[:20]:
                print("  -", e)
        else:
            print("errors: 0")
        print(f"manifest_rows_would_write: {len(manifest_rows)}")
        return 0

    if not manifest_rows:
        print("No manifest rows produced; nothing to write.")
        if errors:
            print(f"Errors: {len(errors)} (first 20)")
            for e in errors[:20]:
                print("  -", e)
        return 3

    df = pd.DataFrame(manifest_rows)
    written_path = _write_manifest_excel(df, output_excel_path)
    print(f"Wrote manifest: {written_path} ({len(df)} rows)")

    if errors:
        print(f"Completed with errors: {len(errors)} (first 20)")
        for e in errors[:20]:
            print("  -", e)
        return 1

    print("Done.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Split audio files into per-segment clips using `segments` column from Excel files in Collected/."
    )
    parser.add_argument(
        "--collected-dir",
        default="Collected",
        help="Directory containing the Excel files and audio folder (default: Collected).",
    )
    parser.add_argument(
        "--output-audio-dir",
        default="Collected/segmented_audio",
        help="Output directory for segment audio clips (default: Collected/segmented_audio).",
    )
    parser.add_argument(
        "--output-excel",
        default="Collected/segmented_dataset.xlsx",
        help="Output Excel manifest path (default: Collected/segmented_dataset.xlsx).",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Inspect Excel structure and segments formats; do not write any files.",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=0,
        help="Process at most N non-empty rows per sheet (0 = no limit). Useful for testing.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print more details (missing headers list, etc).",
    )
    args = parser.parse_args()

    collected_dir = Path(args.collected_dir).resolve()
    output_audio_dir = Path(args.output_audio_dir).resolve()
    output_excel_path = Path(args.output_excel).resolve()

    return run(
        collected_dir=collected_dir,
        output_audio_dir=output_audio_dir,
        output_excel_path=output_excel_path,
        inspect_only=bool(args.inspect),
        max_rows=int(args.max_rows),
        verbose=bool(args.verbose),
    )


if __name__ == "__main__":
    raise SystemExit(main())
